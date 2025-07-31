# items/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.db.models import Q, Count
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from .models import Item, ItemCategory, ItemStatus, ItemTransfer
from .forms import ItemForm, ItemSearchForm, BulkUploadForm, ItemTransferForm, BulkItemSelectForm

class ItemListView(ListView):
    """
    Enhanced ListView for items with search, filtering, and category-wise grouping
    """
    model = Item
    template_name = 'items/item_list.html'
    context_object_name = 'items'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Item.objects.all()
        
        # Get search parameters
        search_query = self.request.GET.get('search', '')
        category_filter = self.request.GET.get('category', '')
        status_filter = self.request.GET.get('status', '')
        
        # Apply search filter
        if search_query:
            queryset = queryset.filter(
                Q(item_name__icontains=search_query) |
                Q(item_id__icontains=search_query) |
                Q(hsn_code__icontains=search_query)
            )
        
        # Apply category filter
        if category_filter:
            queryset = queryset.filter(category=category_filter)
        
        # Apply status filter
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset.order_by('category', 'item_name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add title
        context['title'] = 'Item Management'
        
        # Add search form
        context['search_form'] = ItemSearchForm(self.request.GET)
        
        # Add category-wise statistics
        context['category_stats'] = self.get_category_statistics()
        
        # Add general statistics
        context['total_items'] = Item.objects.count()
        context['active_items'] = Item.objects.filter(status=ItemStatus.ACTIVE).count()
        context['inactive_items'] = Item.objects.filter(status=ItemStatus.INACTIVE).count()
        context['categories_count'] = Item.objects.values('category').distinct().count()
        
        # Preserve search parameters for pagination
        context['search_params'] = self.request.GET.copy()
        if 'page' in context['search_params']:
            del context['search_params']['page']
        
        return context
    
    def get_category_statistics(self):
        """Get item count by category"""
        return Item.objects.values('category').annotate(
            count=Count('id'),
            active_count=Count('id', filter=Q(status=ItemStatus.ACTIVE))
        ).order_by('category')

class ItemDetailView(DetailView):
    """
    Detailed view of a single item
    """
    model = Item
    template_name = 'items/item_detail.html'
    context_object_name = 'item'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add related suppliers (when supplier module is implemented)
        context['item_suppliers'] = self.object.item_suppliers.filter(
            is_active=True
        ).order_by('preference_order')
        
        return context

class ItemCreateView(CreateView):
    """
    Create new item
    """
    model = Item
    form_class = ItemForm
    template_name = 'items/item_form.html'
    success_url = reverse_lazy('items:item_list')
    
    def form_valid(self, form):
        # Set created_by field (when user authentication is implemented)
        form.instance.created_by = getattr(self.request.user, 'username', 'System')
        
        messages.success(
            self.request,
            f'Item "{form.instance.item_name}" has been created successfully!'
        )
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add New Item'
        context['form_action'] = 'Create'
        context['button_text'] = 'Create Item'
        return context

class ItemUpdateView(UpdateView):
    """
    Update existing item
    """
    model = Item
    form_class = ItemForm
    template_name = 'items/item_form.html'
    success_url = reverse_lazy('items:item_list')
    
    def form_valid(self, form):
        # Set updated_by field (when user authentication is implemented)
        form.instance.updated_by = getattr(self.request.user, 'username', 'System')
        
        messages.success(
            self.request,
            f'Item "{form.instance.item_name}" has been updated successfully!'
        )
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Item - {self.object.item_name}'
        context['form_action'] = 'Update'
        context['button_text'] = 'Update Item'
        return context
        context['button_text'] = 'Update Item'
        return context

class ItemDeleteView(DeleteView):
    """
    Delete item with confirmation
    """
    model = Item
    template_name = 'items/item_confirm_delete.html'
    success_url = reverse_lazy('items:item_list')
    
    def delete(self, request, *args, **kwargs):
        item = self.get_object()
        item_name = item.item_name
        
        # Check if item is being used in any purchase orders (future integration)
        # For now, allow deletion but add warning in template
        
        result = super().delete(request, *args, **kwargs)
        
        messages.success(
            request,
            f'Item "{item_name}" has been deleted successfully!'
        )
        return result

# Function-based views for AJAX operations
def item_category_view(request):
    """
    API view to get items by category (for future AJAX calls)
    """
    category = request.GET.get('category')
    if category:
        items = Item.objects.filter(
            category=category,
            status=ItemStatus.ACTIVE
        ).values('id', 'item_name', 'item_id', 'unit')
        return JsonResponse(list(items), safe=False)
    return JsonResponse([], safe=False)

def item_search_api(request):
    """
    API endpoint for item search autocomplete
    """
    query = request.GET.get('q', '')
    if len(query) >= 2:
        items = Item.objects.filter(
            Q(item_name__icontains=query) | Q(item_id__icontains=query),
            status=ItemStatus.ACTIVE
        )[:10]
        
        results = [{
            'id': item.id,
            'text': f"{item.item_id} - {item.item_name}",
            'item_id': item.item_id,
            'category': item.get_category_display(),
            'unit': item.unit
        } for item in items]
        
        return JsonResponse(results, safe=False)
    return JsonResponse([], safe=False)

# Dashboard view for items overview
def items_dashboard(request):
    """
    Dashboard view with item statistics and quick actions
    """
    # Get category statistics
    category_stats = {}
    category_breakdown = Item.objects.values('category').annotate(
        count=Count('id')
    ).order_by('-count')
    
    for category_data in category_breakdown:
        if category_data['category']:
            # Get display name for category
            category_display = dict(ItemCategory.choices).get(category_data['category'], category_data['category'])
            category_stats[category_display] = category_data['count']
    
    context = {
        'title': 'Item Dashboard',
        'total_items': Item.objects.count(),
        'active_items': Item.objects.filter(status=ItemStatus.ACTIVE).count(),
        'inactive_items': Item.objects.filter(status=ItemStatus.INACTIVE).count(),
        'discontinued_items': Item.objects.filter(status=ItemStatus.DISCONTINUED).count(),
        'categories_count': Item.objects.values('category').distinct().count(),
        
        # Category-wise breakdown for charts
        'category_stats': category_stats,
        
        # Recent items
        'recent_items': Item.objects.order_by('-created_at')[:10],
        
        # Items needing attention (low reorder levels)
        'items_needing_attention': Item.objects.filter(
            status=ItemStatus.ACTIVE,
            reorder_level__gt=0
        ).order_by('-reorder_level')[:5],
    }
    
    return render(request, 'items/dashboard.html', context)

# Bulk operations view
def bulk_update_items(request):
    """
    Bulk update items (status, category, etc.)
    """
    if request.method == 'POST':
        item_ids = request.POST.getlist('item_ids')
        action = request.POST.get('action')
        
        if item_ids and action:
            items = Item.objects.filter(id__in=item_ids)
            
            if action == 'activate':
                items.update(status=ItemStatus.ACTIVE)
                messages.success(request, f'{len(item_ids)} items activated successfully!')
            elif action == 'deactivate':
                items.update(status=ItemStatus.INACTIVE)
                messages.success(request, f'{len(item_ids)} items deactivated successfully!')
            elif action == 'discontinue':
                items.update(status=ItemStatus.DISCONTINUED)
                messages.success(request, f'{len(item_ids)} items discontinued successfully!')
        
        return redirect('items:item_list')
    
    return redirect('items:item_list')

# Error handlers
def handler404(request, exception):
    """Custom 404 error handler"""
    return render(request, 'errors/404.html', status=404)

def handler500(request):
    """Custom 500 error handler"""
    return render(request, 'errors/500.html', status=500)
import openpyxl
from django.http import HttpResponse
from django.core.exceptions import ValidationError
import io

# Add this to the existing views.py imports

# Bulk Upload Views
def bulk_upload_view(request):
    """
    Handle bulk upload of items via Excel file
    """
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            
            try:
                # Process the uploaded Excel file
                results = process_excel_file(request, excel_file)
                
                # Store results in session for the results page
                request.session['upload_results'] = results
                request.session['upload_filename'] = excel_file.name
                
                # Redirect to results page
                return redirect('items:bulk_upload_results')
                
            except Exception as e:
                messages.error(
                    request,
                    f'Error processing file: {str(e)}'
                )
    else:
        form = BulkUploadForm()
    
    context = {
        'form': form,
        'title': 'Bulk Upload Items'
    }
    return render(request, 'items/bulk_upload.html', context)

def bulk_upload_results(request):
    """
    Display results of bulk upload operation
    """
    # Get results from session
    results = request.session.get('upload_results')
    filename = request.session.get('upload_filename')
    
    if not results:
        messages.error(request, 'No upload results found. Please try uploading again.')
        return redirect('items:bulk_upload')
    
    # Clear results from session
    if 'upload_results' in request.session:
        del request.session['upload_results']
    if 'upload_filename' in request.session:
        del request.session['upload_filename']
    
    context = {
        'results': results,
        'filename': filename,
        'title': 'Bulk Upload Results'
    }
    return render(request, 'items/bulk_upload_results.html', context)
    
def process_excel_file(request, excel_file):
    """Process the uploaded Excel file and create items using openpyxl"""
    results = {
        'success': 0,
        'errors': 0,
        'error_details': [],
        'success_items': [],  # Track successfully created items
        'skipped_items': [],  # Track skipped items (duplicates)
        'total_rows': 0
    }
    
    try:
        # Read Excel file using openpyxl
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower().replace(' ', '_'))
            else:
                headers.append('')
        
        # Count total rows (excluding header)
        total_rows = ws.max_row - 1
        results['total_rows'] = total_rows
        
        # Required columns
        required_columns = ['item_name', 'category', 'unit', 'status']
        missing_columns = [col for col in required_columns if col not in headers]
        
        if missing_columns:
            raise ValidationError(f'Missing required columns: {", ".join(missing_columns)}')
        
        # Valid choices
        valid_categories = [choice[0] for choice in ItemCategory.choices]
        valid_units = ['kg', 'grams', 'liters', 'ml', 'pcs', 'boxes', 'rolls', 'meters', 'feet', 'bags']
        valid_statuses = [choice[0] for choice in ItemStatus.choices]
        
        # Process each row (starting from row 2)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Skip empty rows
                if not row or not any(row) or not row[0]:
                    continue
                
                # Create a dictionary from headers and row values
                row_data = {}
                for i, header in enumerate(headers):
                    if i < len(row) and header:
                        row_data[header] = row[i]
                
                # Validate required fields
                if not row_data.get('item_name'):
                    continue
                
                item_name = str(row_data['item_name']).strip()
                category = str(row_data.get('category', '')).strip().lower()
                unit = str(row_data.get('unit', '')).strip().lower()
                status = str(row_data.get('status', '')).strip().lower()
                
                # Validate category
                if category not in valid_categories:
                    results['error_details'].append(
                        f'Row {row_num}: Invalid category "{category}". Valid options: {", ".join(valid_categories)}'
                    )
                    results['errors'] += 1
                    continue
                
                # Validate unit
                if unit not in valid_units:
                    results['error_details'].append(
                        f'Row {row_num}: Invalid unit "{unit}". Valid options: {", ".join(valid_units)}'
                    )
                    results['errors'] += 1
                    continue
                
                # Validate status
                if status not in valid_statuses:
                    results['error_details'].append(
                        f'Row {row_num}: Invalid status "{status}". Valid options: {", ".join(valid_statuses)}'
                    )
                    results['errors'] += 1
                    continue
                
                # Check if item already exists
                if Item.objects.filter(item_name__iexact=item_name).exists():
                    results['skipped_items'].append(f'Row {row_num}: "{item_name}" (already exists)')
                    results['errors'] += 1
                    continue
                
                # Prepare item data
                item_data = {
                    'item_name': item_name,
                    'category': category,
                    'unit': unit,
                    'status': status,
                    'created_by': getattr(request.user, 'username', 'Bulk Upload')
                }
                
                # Add optional fields if present
                optional_fields = {
                    'hsn_code': str,
                    'tax_rate': float,
                    'reorder_level': int,
                    'standard_rate': float,
                    'shelf_life_days': int,
                    'specification': str,
                    'description': str,
                    'remarks': str
                }
                
                for field_name, field_type in optional_fields.items():
                    if field_name in row_data and row_data[field_name] is not None:
                        value = row_data[field_name]
                        if value != '' and str(value).strip() != '':
                            try:
                                if field_type in [float, int]:
                                    value = field_type(value)
                                else:
                                    value = str(value).strip()
                                item_data[field_name] = value
                            except (ValueError, TypeError):
                                results['error_details'].append(
                                    f'Row {row_num}: Invalid {field_name} value "{value}"'
                                )
                                continue
                
                # Create the item
                item = Item.objects.create(**item_data)
                results['success'] += 1
                results['success_items'].append(f'Row {row_num}: "{item_name}" ({item.item_id})')
                
            except Exception as e:
                results['error_details'].append(
                    f'Row {row_num}: {str(e)}'
                )
                results['errors'] += 1
        
    except Exception as e:
        raise ValidationError(f'Error reading Excel file: {str(e)}')
    
    return results

def download_template(request):
    """
    Download Excel template for bulk upload with data validation
    """
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items Template"
    
    # Define headers
    headers = [
        'item_name', 'category', 'unit', 'status', 'hsn_code', 
        'tax_rate', 'reorder_level', 'standard_rate', 'shelf_life_days',
        'specification', 'description', 'remarks'
    ]
    
    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    
    # Add sample data rows
    sample_data = [
        [
            'Basmati Rice Premium', 'raw_material', 'kg', 'active', '1006',
            5.0, 1000, 85.50, 365, 'Premium quality long grain', 
            'High quality Basmati rice', 'Sample item'
        ],
        [
            'Cardboard Boxes', 'packaging', 'pcs', 'active', '4819',
            18.0, 500, 12.50, '', '3-ply corrugated boxes',
            'Standard packaging boxes', ''
        ],
        [
            'Office Paper A4', 'stationery', 'boxes', 'active', '4802',
            18.0, 50, 285.00, '', 'Premium quality paper',
            'Office stationery paper', ''
        ]
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Add data validation for dropdowns
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Category validation
    category_values = ['raw_material', 'packaging', 'stationery', 'infrastructure', 'others']
    category_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(category_values)}"',
        showDropDown=True
    )
    category_validation.error = "Invalid category! Choose from: " + ", ".join(category_values)
    category_validation.errorTitle = "Invalid Category"
    ws.add_data_validation(category_validation)
    category_validation.add(f'B2:B1000')  # Apply to category column
    
    # Unit validation  
    unit_values = ['kg', 'grams', 'liters', 'ml', 'pcs', 'boxes', 'rolls', 'meters', 'feet', 'bags']
    unit_validation = DataValidation(
        type="list", 
        formula1=f'"{",".join(unit_values)}"',
        showDropDown=True
    )
    unit_validation.error = "Invalid unit! Choose from: " + ", ".join(unit_values)
    unit_validation.errorTitle = "Invalid Unit"
    ws.add_data_validation(unit_validation)
    unit_validation.add(f'C2:C1000')  # Apply to unit column
    
    # Status validation
    status_values = ['active', 'inactive', 'discontinued']
    status_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(status_values)}"', 
        showDropDown=True
    )
    status_validation.error = "Invalid status! Choose from: " + ", ".join(status_values)
    status_validation.errorTitle = "Invalid Status"
    ws.add_data_validation(status_validation)
    status_validation.add(f'D2:D1000')  # Apply to status column
    
    # Add validation information in a separate sheet
    validation_sheet = wb.create_sheet("Instructions & Valid Values")
    
    # Instructions and validation info
    instructions = [
        ['BULK UPLOAD INSTRUCTIONS', '', '', ''],
        ['', '', '', ''],
        ['REQUIRED COLUMNS (Must be filled):', '', '', ''],
        ['Column Name', 'Required', 'Valid Values', 'Example'],
        ['item_name', 'YES', 'Any unique text', 'Basmati Rice Premium'],
        ['category', 'YES', 'raw_material, packaging, stationery, infrastructure, others', 'raw_material'],
        ['unit', 'YES', 'kg, grams, liters, ml, pcs, boxes, rolls, meters, feet, bags', 'kg'],
        ['status', 'YES', 'active, inactive, discontinued', 'active'],
        ['', '', '', ''],
        ['OPTIONAL COLUMNS (Can be left empty):', '', '', ''],
        ['hsn_code', 'NO', '4-8 digit HSN/SAC code', '1006'],
        ['tax_rate', 'NO', 'Number (0-100)', '18.0'],
        ['reorder_level', 'NO', 'Positive number', '1000'],
        ['standard_rate', 'NO', 'Positive number', '85.50'],
        ['shelf_life_days', 'NO', 'Positive number', '365'],
        ['specification', 'NO', 'Any text', 'Premium quality'],
        ['description', 'NO', 'Any text', 'Detailed description'],
        ['remarks', 'NO', 'Any text', 'Additional notes'],
        ['', '', '', ''],
        ['CATEGORY OPTIONS:', '', '', ''],
        ['raw_material', 'For food ingredients, spices, etc.', '', ''],
        ['packaging', 'For boxes, bags, wrapping materials', '', ''],
        ['stationery', 'For office supplies, paper, etc.', '', ''],
        ['infrastructure', 'For equipment, machinery, etc.', '', ''],
        ['others', 'For items not in above categories', '', ''],
        ['', '', '', ''],
        ['UNIT OPTIONS:', '', '', ''],
        ['kg', 'Kilogram (weight)', '', ''],
        ['grams', 'Grams (weight)', '', ''],
        ['liters', 'Liters (volume)', '', ''],
        ['ml', 'Milliliters (volume)', '', ''],
        ['pcs', 'Pieces (count)', '', ''],
        ['boxes', 'Boxes (count)', '', ''],
        ['rolls', 'Rolls (count)', '', ''],
        ['meters', 'Meters (length)', '', ''],
        ['feet', 'Feet (length)', '', ''],
        ['bags', 'Bags (count)', '', ''],
        ['', '', '', ''],
        ['STATUS OPTIONS:', '', '', ''],
        ['active', 'Item is currently in use', '', ''],
        ['inactive', 'Item is temporarily not in use', '', ''],
        ['discontinued', 'Item is no longer used', '', ''],
        ['', '', '', ''],
        ['IMPORTANT TIPS:', '', '', ''],
        ['• Item names must be unique', '', '', ''],
        ['• Use exact values from dropdown lists', '', '', ''],
        ['• Numbers should not contain commas or symbols', '', '', ''],
        ['• Leave optional fields empty if not needed', '', '', ''],
        ['• Maximum file size: 5MB', '', '', ''],
        ['• Supported formats: .xlsx, .xls', '', '', ''],
    ]
    
    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = validation_sheet.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:  # Title row
                cell.font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            elif row_num in [3, 10, 20, 28, 35, 43]:  # Section headers
                cell.font = openpyxl.styles.Font(bold=True, color="366092")
            elif row_num == 4:  # Column headers
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Adjust column widths for both sheets
    for sheet in [ws, validation_sheet]:
        for column_cells in sheet.columns:
            length = max(len(str(cell.value) or '') for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="items_bulk_upload_template.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    return response

# Error handlers
def handler404(request, exception):
    """Custom 404 error handler"""
    return render(request, 'errors/404.html', status=404)

def handler500(request):
    """Custom 500 error handler"""
    return render(request, 'errors/500.html', status=500)


# Transfer Widget Views
def item_transfer_widget(request):
    """
    Widget for transferring individual items
    """
    if request.method == 'POST':
        form = ItemTransferForm(request.POST)
        if form.is_valid():
            # Create transfer record
            transfer = ItemTransfer.objects.create(
                item=form.cleaned_data['item'],
                quantity=form.cleaned_data['quantity'],
                from_location=form.cleaned_data['from_location'],
                to_location=form.cleaned_data['to_location'],
                transfer_reason=form.cleaned_data['transfer_reason'],
                remarks=form.cleaned_data['remarks'],
                created_by=request.user.username if hasattr(request, 'user') and request.user.is_authenticated else 'Anonymous'
            )
            
            messages.success(
                request,
                f"Transfer {transfer.transfer_id} created successfully! "
                f"{transfer.quantity} {transfer.item.unit} of {transfer.item.item_name} "
                f"from {transfer.from_location} to {transfer.to_location}."
            )
            
            return redirect('items:transfer_list')
    else:
        form = ItemTransferForm()
    
    return render(request, 'items/transfer_widget.html', {
        'form': form,
        'title': 'Item Transfer Widget'
    })

def transfer_list(request):
    """
    List all transfers with filtering and search
    """
    transfers = ItemTransfer.objects.select_related('item').all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        transfers = transfers.filter(
            Q(transfer_id__icontains=search_query) |
            Q(item__item_name__icontains=search_query) |
            Q(from_location__icontains=search_query) |
            Q(to_location__icontains=search_query)
        )
    
    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        transfers = transfers.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(transfers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'transfers': page_obj,
        'title': 'Transfer History',
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': ItemTransfer.TRANSFER_STATUS_CHOICES,
        'total_transfers': transfers.count(),
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
    }
    
    return render(request, 'items/transfer_list.html', context)

def bulk_item_select(request):
    """
    Bulk item selection for various operations
    """
    if request.method == 'POST':
        form = BulkItemSelectForm(request.POST)
        if form.is_valid():
            selected_items = form.cleaned_data['selected_items']
            action = form.cleaned_data['action']
            
            if action == 'transfer':
                # Store selected items in session for transfer form
                request.session['bulk_transfer_items'] = [item.id for item in selected_items]
                return redirect('items:bulk_transfer')
            
            elif action == 'activate':
                updated_count = selected_items.update(status='active')
                messages.success(request, f"Successfully activated {updated_count} items.")
                
            elif action == 'deactivate':
                updated_count = selected_items.update(status='inactive')
                messages.success(request, f"Successfully deactivated {updated_count} items.")
                
            elif action == 'export':
                # TODO: Implement CSV export
                messages.info(request, "Export functionality coming soon!")
            
            return redirect('items:item_list')
    else:
        form = BulkItemSelectForm()
    
    return render(request, 'items/bulk_select.html', {
        'form': form,
        'title': 'Bulk Item Operations'
    })

def bulk_transfer(request):
    """
    Handle bulk transfer of selected items
    """
    # Get selected items from session
    item_ids = request.session.get('bulk_transfer_items', [])
    if not item_ids:
        messages.error(request, "No items selected for transfer.")
        return redirect('items:bulk_select')
    
    selected_items = Item.objects.filter(id__in=item_ids)
    
    if request.method == 'POST':
        # Process bulk transfer
        from_location = request.POST.get('from_location')
        to_location = request.POST.get('to_location')
        transfer_reason = request.POST.get('transfer_reason', '')
        
        if from_location and to_location:
            transfers_created = 0
            for item in selected_items:
                # Default quantity of 1 for bulk transfers (can be customized)
                quantity = request.POST.get(f'quantity_{item.id}', 1)
                try:
                    quantity = float(quantity)
                    if quantity > 0:
                        ItemTransfer.objects.create(
                            item=item,
                            quantity=quantity,
                            from_location=from_location,
                            to_location=to_location,
                            transfer_reason=transfer_reason,
                            created_by=request.user.username if hasattr(request, 'user') and request.user.is_authenticated else 'Anonymous'
                        )
                        transfers_created += 1
                except (ValueError, TypeError):
                    continue
            
            if transfers_created > 0:
                messages.success(
                    request,
                    f"Successfully created {transfers_created} transfer records from "
                    f"{from_location} to {to_location}."
                )
            else:
                messages.error(request, "No valid transfers could be created.")
            
            # Clear session data
            if 'bulk_transfer_items' in request.session:
                del request.session['bulk_transfer_items']
            
            return redirect('items:transfer_list')
        else:
            messages.error(request, "Both from and to locations are required.")
    
    context = {
        'selected_items': selected_items,
        'title': 'Bulk Transfer Items',
        'item_count': selected_items.count()
    }
    
    return render(request, 'items/bulk_transfer.html', context)
